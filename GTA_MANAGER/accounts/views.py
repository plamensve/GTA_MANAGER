import datetime

import openpyxl
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django.views.generic import TemplateView
from django.contrib.auth import login
from openpyxl.styles import Font, Alignment, Border, Side

from .forms import CustomUserCreationForm, CustomAuthenticationForm, VehicleCreateForm, CustomUserChangeForm, \
    VehicleFullDetailsCreateForm, VehicleEditForm
from .models import Vehicles, VehicleFullDetails
from .utils import get_all_vehicles, vehicle_full_details_info
from openpyxl import Workbook

from ..web.utils import check_expiration, send_email_registration


class IndexView(TemplateView):
    template_name = 'index.html'

    @method_decorator(never_cache)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.method == 'POST':
            register_form = CustomUserCreationForm(self.request.POST)
            form_invalid = not register_form.is_valid()
        else:
            register_form = CustomUserCreationForm()
            form_invalid = False

        login_form = CustomAuthenticationForm()

        context['register_form'] = register_form
        context['login_form'] = login_form
        context['form_invalid'] = form_invalid
        return context

    def post(self, request, *args, **kwargs):
        context = self.get_context_data()

        if 'register' in request.POST:
            form = CustomUserCreationForm(request.POST)
            if form.is_valid():
                user = form.save()
                login(request, user)
                return redirect('after-register')
            else:
                context['register_form'] = form  # Връща формата с грешки обратно в контекста

        elif 'login' in request.POST:
            form = CustomAuthenticationForm(data=request.POST)
            if form.is_valid():
                login(request, form.get_user())
                return redirect('front-page')
            else:
                context['login_form'] = form  # Връща формата с грешки обратно в контекста

        return self.render_to_response(context)


@login_required(login_url='index')
def front_page(request):
    all_vehicles = get_all_vehicles()
    current_user = request.user

    contex = {
        'all_vehicles': all_vehicles,
        'current_user': current_user,
    }

    if current_user.is_authenticated:
        return render(request, 'front_page.html', contex)


@login_required(login_url='index')
def profile_page(request):
    current_user = request.user

    contex = {
        'current_user': current_user
    }

    return render(request, 'profile-page.html', contex)


@login_required(login_url='index')
def edit_profile_page(request):
    current_user = request.user

    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, instance=current_user)
        if form.is_valid():
            form.save()
            return redirect('profile-page')  # Пренасочване след редакция
    else:
        form = CustomUserChangeForm(instance=current_user)  # Зарежда текущите данни на потребителя

    context = {
        'form': form,
        'current_user': current_user
    }

    return render(request, 'edit-profile-page.html', context)


@login_required(login_url='index')
def add_vehicle(request):
    form_vehicle = VehicleCreateForm()

    if request.method == 'POST':
        form_vehicle = VehicleCreateForm(request.POST)
        if form_vehicle.is_valid():
            form_vehicle.save()
            return redirect('front-page')

    context = {
        'form_vehicle': form_vehicle
    }
    return render(request, 'vehicles/add-vehicles.html', context)


@login_required(login_url='index')
def vehicle_details(request, pk):
    vehicle_details = vehicle_full_details_info(pk)
    vehicle_information = VehicleFullDetails.objects.filter(vehicle_id=pk).first()

    try:
        current_vehicle = Vehicles.objects.get(pk=pk)
    except Vehicles.DoesNotExist:
        return render(request, '404.html', status=404)

    condition_class = 'status-active' if current_vehicle.condition == 'АКТИВЕН' else 'status-inactive'

    context = {
        'current_vehicle': current_vehicle,
        'condition_class': condition_class,
        'vehicle_details': vehicle_details,
        'vehicle_information': vehicle_information if vehicle_information else None,
    }

    return render(request, 'vehicles/vehicle-details.html', context)


@login_required(login_url='index')
def add_full_information(request, pk):
    vehicle = get_object_or_404(Vehicles, pk=pk)

    if request.method == 'POST':
        form = VehicleFullDetailsCreateForm(request.POST)
        if form.is_valid():
            vehicle_full_details = form.save(commit=False)
            vehicle_full_details.vehicle = vehicle
            vehicle_full_details.save()
            return redirect('vehicle_details', pk=pk)
    else:
        form = VehicleFullDetailsCreateForm()

    context = {
        'form': form,
    }

    return render(request, 'vehicles/add-full-information.html', context)


@login_required(login_url='index')
def edit_full_information(request, pk):
    vehicle_full_details = get_object_or_404(Vehicles, pk=pk)
    vehicle_second_full_details = get_object_or_404(VehicleFullDetails, vehicle_id=pk)

    if request.method == 'POST':
        form = VehicleEditForm(request.POST, instance=vehicle_full_details)
        form_2 = VehicleFullDetailsCreateForm(request.POST, instance=vehicle_second_full_details)

        if form.is_valid() and form_2.is_valid():
            # Save the first form
            form.save()

            # Save the second form with a ForeignKey to vehicle_full_details
            second_details = form_2.save(commit=False)
            second_details.vehicle = vehicle_full_details  # Свързване на ForeignKey
            second_details.save()

            return redirect('vehicle_details', pk=pk)
    else:
        form = VehicleEditForm(instance=vehicle_full_details)
        form_2 = VehicleFullDetailsCreateForm(instance=vehicle_second_full_details)

    context = {
        'form': form,
        'form_2': form_2,
        'vehicle_full_details': vehicle_full_details,
    }

    return render(request, 'vehicles/edit-full-information.html', context)


@login_required(login_url='index')
def delete_information(request, pk):
    vehicle_full_details = get_object_or_404(Vehicles, pk=pk)

    if request.method == 'POST':
        vehicle_full_details.delete()
        return redirect('front-page')

    # Ако е GET заявка, покажете страницата за потвърждение
    return render(request, 'vehicles/delete-information.html', {'vehicle': vehicle_full_details})


@login_required(login_url='index')
def generate_vehicle_report(request):
    # Създаване на Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Служебни превозни средства"

    # Стилове
    bold_font = Font(bold=True, size=14)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Добавяне на заглавия
    header_info = f'ДЖИ ТИ ЕЙ ПЕТРОЛИУМ ООД - АВТОМОБИЛИ / ВЛЕКАЧИ / ЦИСТЕРНИ <<{datetime.datetime.now().strftime('%d-%m-%Y')}>>'
    ws.append([header_info])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)  # Обединяване на клетките
    header_cell = ws.cell(row=1, column=1)
    header_cell.font = Font(bold=True, size=16)
    header_cell.alignment = center_alignment

    headers = ["№", "Тип", "Марка", "Модел", "Рег. номер", "Състояние"]
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border

    # Добавяне на данни от базата
    all_vehicles = Vehicles.objects.all()
    for idx, vehicle in enumerate(all_vehicles, start=1):
        row = [
            idx,
            vehicle.type.upper(),
            vehicle.brand.upper(),
            vehicle.model.upper(),
            vehicle.register_number.upper(),
            vehicle.condition.upper()
        ]
        ws.append(row)

        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=idx + 2, column=col_num)
            cell.alignment = center_alignment
            cell.border = border

    # Настройки за ширина на колоните
    column_widths = [10, 20, 25, 25, 30, 20]  # Увеличени ширини
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Настройки за отговор
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Fleet_Masters_Report.xlsx"'

    # Запазване на Excel файла в отговора
    wb.save(response)
    return response


def send_email(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        if not email:
            return JsonResponse({'error': 'Email is required'}, status=400)

        # Вашата логика за изпращане на имейл
        expiration_documents = check_expiration()

        documents_list = [
            {
                'current_registration_number': docs['current_registration_number'],
                'current_type_name': docs['current_type_name'],
                'insurance_civil_liability': docs['insurance_civil_liability'],
                'insurance_casco_validity': docs['insurance_casco_validity'],
                'tachograph_validity': docs['tachograph_validity'],
                'adr_validity': docs['adr_validity'],
                'fitness_protocol_validity': docs['fitness_protocol_validity'],
                'technical_check_validity': docs['technical_check_validity'],
            }
            for docs in expiration_documents
        ]

        html_message = render_to_string('email_templates/expiration_email.html', {
            'documents_list': documents_list,
        })

        send_mail(
            subject="Обобщена информация за изтичащи документи",
            message=None,
            from_email='svetoslavov.plamen@gmail.com',
            recipient_list=[email],
            html_message=html_message,
        )
        return JsonResponse({'success': 'Email sent successfully'})
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required(login_url='index')
def generate_vehicle_report_info(request):
    # Създаване на Excel файл
    wb = Workbook()
    ws = wb.active
    ws.title = "Документи"

    # Стилове
    bold_font = Font(bold=True, size=16)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Добавяне на заглавия
    header_info = 'ДЖИ ТИ ЕЙ ПЕТРОЛИУМ ООД - Придружаващи документи'
    ws.append([header_info])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)  # Обединяване на клетките
    header_cell = ws.cell(row=1, column=1)
    header_cell.font = bold_font
    header_cell.alignment = center_alignment
    header_cell.border = border

    headers = ["№", "Рег. №", "Гражданска отговорност", "Каско", "Тахограф",
               "АДР", "Протокол за годност", "Технически преглед"]
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border

    all_vehicles = VehicleFullDetails.objects.select_related('vehicle')

    for idx, vehicle_detail in enumerate(all_vehicles, start=1):
        row = [
            idx,
            vehicle_detail.vehicle.register_number or 'N/A',

            vehicle_detail.insurance_civil_liability.strftime('%d-%m-%Y')
            if vehicle_detail.insurance_civil_liability else 'N/A',

            vehicle_detail.insurance_casco_validity.strftime('%d-%m-%Y')
            if vehicle_detail.insurance_casco_validity else 'N/A',

            vehicle_detail.tachograph_validity.strftime('%d-%m-%Y')
            if vehicle_detail.tachograph_validity else 'N/A',

            vehicle_detail.adr_validity.strftime('%d-%m-%Y')
            if vehicle_detail.adr_validity else 'N/A',

            vehicle_detail.fitness_protocol_validity.strftime('%d-%m-%Y')
            if vehicle_detail.fitness_protocol_validity else 'N/A',

            vehicle_detail.technical_check_validity.strftime('%d-%m-%Y')
            if vehicle_detail.technical_check_validity else 'N/A',
        ]
        # Добавете реда към вашия Excel файл или друг изход
        ws.append(row)

        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=idx + 2, column=col_num)
            cell.alignment = center_alignment
            cell.border = border

    # Индивидуална ширина на колоните
    column_widths = [5.0, 20.0, 30.0, 30.0, 25.0, 25.0, 30.0, 30.0]  # Индивидуални ширини за всяка колона
    for col_num, width in enumerate(column_widths, start=1):
        col_letter = ws.cell(row=2, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = width

    # Настройки за отговор
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Expiring_Documents_Tracker.xlsx"'

    # Запазване на Excel файла в отговора
    wb.save(response)
    return response


def freight_trains(request):
    trucks = Vehicles.objects.filter(type='ВЛЕКАЧ')
    tanks = Vehicles.objects.filter(type='ЦИСТЕРНА')
    compositions = []
    insurance_civil_liability = []
    casco_validity = []
    tachograph_validity = []
    adr = []
    fitness_protocol_validity = []
    technical_check_validity = []

    for tr in trucks:
        for tn in tanks:
            if tn.comp == tr.comp:  # Проверка за съвпадение
                compositions.append(f"{tr.register_number} / {tn.register_number}")

                truck_insurance = VehicleFullDetails.objects.get(vehicle_id=tr.pk)
                tank_insurance = VehicleFullDetails.objects.get(vehicle_id=tn.pk)
                insurance_civil_liability.append(
                    f"{truck_insurance.insurance_civil_liability.strftime('%d-%m-%Y') if truck_insurance.insurance_civil_liability else 'N/A'} / "
                    f"{tank_insurance.insurance_civil_liability.strftime('%d-%m-%Y') if tank_insurance.insurance_civil_liability else 'N/A'}"
                )

                truck_casco = VehicleFullDetails.objects.get(vehicle_id=tr.pk)
                tank_casco = VehicleFullDetails.objects.get(vehicle_id=tn.pk)
                casco_validity.append(
                    f"{truck_casco.insurance_casco_validity.strftime('%d-%m-%Y') if truck_casco.insurance_casco_validity else 'N/A'} / "
                    f"{tank_casco.insurance_casco_validity.strftime('%d-%m-%Y') if tank_casco.insurance_casco_validity else 'N/A'}"
                )

                truck_tachograph_validity = VehicleFullDetails.objects.get(vehicle_id=tr.pk)
                tank_tachograph_validity = VehicleFullDetails.objects.get(vehicle_id=tn.pk)
                tachograph_validity.append(
                    f"{truck_tachograph_validity.tachograph_validity.strftime('%d-%m-%Y') if truck_tachograph_validity.tachograph_validity else 'N/A'} / "
                    f"{tank_tachograph_validity.tachograph_validity.strftime('%d-%m-%Y') if tank_tachograph_validity.tachograph_validity else 'N/A'}"
                )

                truck_adr = VehicleFullDetails.objects.get(vehicle_id=tr.pk)
                tank_adr = VehicleFullDetails.objects.get(vehicle_id=tn.pk)
                adr.append(
                    f"{truck_adr.adr_validity.strftime('%d-%m-%Y') if truck_adr.adr_validity else 'N/A'} / "
                    f"{tank_adr.adr_validity.strftime('%d-%m-%Y') if tank_adr.adr_validity else 'N/A'}"
                )

                truck_fitness = VehicleFullDetails.objects.get(vehicle_id=tr.pk)
                tank_fitness = VehicleFullDetails.objects.get(vehicle_id=tn.pk)
                fitness_protocol_validity.append(
                    f"{truck_fitness.fitness_protocol_validity.strftime('%d-%m-%Y') if truck_fitness.fitness_protocol_validity else 'N/A'} / "
                    f"{tank_fitness.fitness_protocol_validity.strftime('%d-%m-%Y') if tank_fitness.fitness_protocol_validity else 'N/A'}"
                )

                truck_technical_check = VehicleFullDetails.objects.get(vehicle_id=tr.pk)
                tank_technical_check = VehicleFullDetails.objects.get(vehicle_id=tn.pk)
                technical_check_validity.append(
                    f"{truck_technical_check.technical_check_validity.strftime('%d-%m-%Y') if truck_technical_check.technical_check_validity else 'N/A'} / "
                    f"{tank_technical_check.technical_check_validity.strftime('%d-%m-%Y') if tank_technical_check.technical_check_validity else 'N/A'}"
                )

    context = {
        'trucks': trucks,
        'tanks': tanks,
        'compositions': compositions,
        'insurance_civil_liability': insurance_civil_liability,
        'casco_validity': casco_validity,
        'tachograph_validity': tachograph_validity,
        'adr': adr,
        'fitness_protocol_validity': fitness_protocol_validity,
        'technical_check_validity': technical_check_validity,
    }

    return render(request, 'vehicles/freight-trains.html', context)


@login_required(login_url='index')
def export_to_excel(request):
    # Създаване на нов Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Документи"

    # Стилове
    bold_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Заглавие на таблицата
    header_info = f"ДЖИ ТИ ЕЙ ПЕТРОЛИУМ ООД - Придружаващи документи <<{datetime.datetime.now().strftime('%d-%m-%Y')}>>"
    ws.append([header_info])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    header_cell = ws.cell(row=1, column=1)
    header_cell.font = Font(bold=True, size=16)
    header_cell.alignment = center_alignment

    # Заглавия на колоните
    headers = [
        "№", "Състав", "Гражданска отговорност", "Каско",
        "Тахограф", "АДР", "Протокол за годност", "Технически преглед"
    ]
    ws.append(headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num)
        cell.font = bold_font
        cell.alignment = center_alignment
        cell.border = border

    # Данни за таблицата
    trucks = Vehicles.objects.filter(type='ВЛЕКАЧ')
    tanks = Vehicles.objects.filter(type='ЦИСТЕРНА')
    row_index = 3

    for idx, tr in enumerate(trucks, start=1):
        for tn in tanks:
            if tn.comp == tr.comp:  # Проверка за съвпадение
                composition = f"{tr.register_number} / {tn.register_number}"

                truck_detail = VehicleFullDetails.objects.get(vehicle_id=tr.pk)
                tank_detail = VehicleFullDetails.objects.get(vehicle_id=tn.pk)

                row = [
                    idx,
                    composition,
                    f"{truck_detail.insurance_civil_liability.strftime('%d-%m-%Y') if truck_detail.insurance_civil_liability else 'N/A'} / "
                    f"{tank_detail.insurance_civil_liability.strftime('%d-%m-%Y') if tank_detail.insurance_civil_liability else 'N/A'}",
                    f"{truck_detail.insurance_casco_validity.strftime('%d-%m-%Y') if truck_detail.insurance_casco_validity else 'N/A'} / "
                    f"{tank_detail.insurance_casco_validity.strftime('%d-%m-%Y') if tank_detail.insurance_casco_validity else 'N/A'}",
                    f"{truck_detail.tachograph_validity.strftime('%d-%m-%Y') if truck_detail.tachograph_validity else 'N/A'} / "
                    f"{tank_detail.tachograph_validity.strftime('%d-%m-%Y') if tank_detail.tachograph_validity else 'N/A'}",
                    f"{truck_detail.adr_validity.strftime('%d-%m-%Y') if truck_detail.adr_validity else 'N/A'} / "
                    f"{tank_detail.adr_validity.strftime('%d-%m-%Y') if tank_detail.adr_validity else 'N/A'}",
                    f"{truck_detail.fitness_protocol_validity.strftime('%d-%m-%Y') if truck_detail.fitness_protocol_validity else 'N/A'} / "
                    f"{tank_detail.fitness_protocol_validity.strftime('%d-%m-%Y') if tank_detail.fitness_protocol_validity else 'N/A'}",
                    f"{truck_detail.technical_check_validity.strftime('%d-%m-%Y') if truck_detail.technical_check_validity else 'N/A'} / "
                    f"{tank_detail.technical_check_validity.strftime('%d-%m-%Y') if tank_detail.technical_check_validity else 'N/A'}",
                ]

                ws.append(row)

                for col_num, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_index, column=col_num)
                    cell.alignment = center_alignment
                    cell.border = border

                row_index += 1

    # Индивидуална ширина на колоните
    column_widths = [5, 25, 30, 30, 25, 25, 30, 30]
    for col_num, width in enumerate(column_widths, start=1):
        col_letter = ws.cell(row=2, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = width

    # Отговора като Excel файл
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Freight_Trains_Report.xlsx"'
    wb.save(response)
    return response


def after_register(request):
    return render(request, 'after-register.html')
